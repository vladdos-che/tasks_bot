from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from typing import List
from datetime import date, timedelta
from fastapi.responses import StreamingResponse
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import calendar
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import numpy as np

from db.database import get_db
from db.models import Schedule, ScheduleAssignment, AdminUser, RoleTaskMatrix, user_roles, User, Task
from api import schemas
from api.core.deps import get_current_admin_user

router = APIRouter()

@router.get("", response_model=List[schemas.ScheduleOut])
async def get_schedules(
    db: AsyncSession = Depends(get_db),
    current_admin: AdminUser = Depends(get_current_admin_user)
):
    result = await db.execute(select(Schedule).options(selectinload(Schedule.assignments)))
    return result.scalars().all()

@router.post("", response_model=schemas.ScheduleOut)
async def create_schedule(
    schedule_in: schemas.ScheduleCreate,
    db: AsyncSession = Depends(get_db),
    current_admin: AdminUser = Depends(get_current_admin_user)
):
    # Check if schedule exists for this date. If so, delete it to replace.
    existing = await db.execute(select(Schedule).where(Schedule.date == schedule_in.date))
    existing_schedule = existing.scalars().first()
    if existing_schedule:
        await db.delete(existing_schedule)
        await db.flush()
        
    # Verify users have permissions for the assigned tasks
    for assignment in schedule_in.assignments:
        if assignment.main_user_id:
            # Check if user has any role assigned to this task
            user_roles_result = await db.execute(select(user_roles.c.role_id).where(user_roles.c.user_id == assignment.main_user_id))
            user_role_ids = user_roles_result.scalars().all()
            
            if not user_role_ids:
                raise HTTPException(status_code=400, detail=f"User {assignment.main_user_id} has no roles assigned")

            perm = await db.execute(
                select(RoleTaskMatrix)
                .where(RoleTaskMatrix.role_id.in_(user_role_ids))
                .where(RoleTaskMatrix.task_id == assignment.task_id)
            )
            if not perm.scalars().first():
                raise HTTPException(status_code=400, detail=f"User {assignment.main_user_id} does not have permission for task {assignment.task_id}")
                
    new_schedule = Schedule(date=schedule_in.date, meeting_type=schedule_in.meeting_type)
    db.add(new_schedule)
    await db.flush() # To get new_schedule.id
    
    for assignment in schedule_in.assignments:
        new_assignment = ScheduleAssignment(
            schedule_id=new_schedule.id,
            task_id=assignment.task_id,
            main_user_id=assignment.main_user_id,
            helper_user_id=assignment.helper_user_id,
            custom_name=assignment.custom_name
        )
        db.add(new_assignment)
        
    await db.commit()
    await db.refresh(new_schedule)
    
    # Needs to eagerly load assignments to return
    result = await db.execute(select(Schedule).where(Schedule.id == new_schedule.id).options(selectinload(Schedule.assignments)))
    return result.scalars().first()

@router.delete("/{schedule_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_schedule(
    schedule_id: int,
    db: AsyncSession = Depends(get_db),
    current_admin: AdminUser = Depends(get_current_admin_user)
):
    result = await db.execute(select(Schedule).where(Schedule.id == schedule_id))
    schedule = result.scalars().first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    await db.delete(schedule)
    await db.commit()

@router.post("/auto-draft", response_model=List[schemas.ScheduleAssignmentCreate])
async def auto_draft_schedule(
    request: schemas.AutoDraftRequest,
    db: AsyncSession = Depends(get_db),
    current_admin: AdminUser = Depends(get_current_admin_user)
):
    # 1. Fetch tasks
    if request.meeting_type == 'Будни':
        tasks_result = await db.execute(select(Task).where(Task.is_weekday == True))
    else:
        tasks_result = await db.execute(select(Task).where(Task.is_weekend == True))
    
    all_tasks = tasks_result.scalars().all()
    
    # Filter out excluded tasks
    excluded_ids = []
    if request.exclude_tasks:
        try:
            excluded_ids = [int(t_id) for t_id in request.exclude_tasks.split(',') if t_id.strip()]
        except ValueError:
            pass
            
    tasks = [t for t in all_tasks if t.id not in excluded_ids]

    # 2. Fetch active users and their roles
    users_result = await db.execute(select(User).where(User.is_active == True).options(selectinload(User.roles)))
    active_users = users_result.scalars().all()

    # 3. Fetch role-task matrix
    matrix_result = await db.execute(select(RoleTaskMatrix))
    matrix = matrix_result.scalars().all()

    # 4. Sort tasks by restrictiveness (fewest eligible brothers first)
    def get_eligible_count(task):
        allowed_roles = {m.role_id for m in matrix if m.task_id == task.id}
        return sum(1 for u in active_users if any(r.id in allowed_roles for r in u.roles))
        
    tasks.sort(key=get_eligible_count)

    # 5. Calculate last assignment date and assignment count for each user
    assignments_result = await db.execute(
        select(ScheduleAssignment.main_user_id, ScheduleAssignment.helper_user_id, Schedule.date, ScheduleAssignment.task_id)
        .join(Schedule, Schedule.id == ScheduleAssignment.schedule_id)
    )
    last_dates = {}
    assignment_counts = {}
    task_assignment_counts = {} # user_id -> { task_id -> count }
    min_distances = {} # user_id -> min absolute days
    
    today = date.today()
    draft_date = date.fromisoformat(request.date)
    window_start = today - timedelta(days=30)
    window_end = today + timedelta(days=60)
    
    for main_user_id, helper_user_id, assignment_date, task_id in assignments_result.all():
        dist = abs((assignment_date - draft_date).days)
        if main_user_id:
            if main_user_id not in last_dates or assignment_date > last_dates[main_user_id]:
                last_dates[main_user_id] = assignment_date
            if main_user_id not in min_distances or dist < min_distances[main_user_id]:
                min_distances[main_user_id] = dist
            if window_start <= assignment_date <= window_end:
                assignment_counts[main_user_id] = assignment_counts.get(main_user_id, 0) + 1
                if main_user_id not in task_assignment_counts:
                    task_assignment_counts[main_user_id] = {}
                task_assignment_counts[main_user_id][task_id] = task_assignment_counts[main_user_id].get(task_id, 0) + 1
        if helper_user_id:
            if helper_user_id not in last_dates or assignment_date > last_dates[helper_user_id]:
                last_dates[helper_user_id] = assignment_date
            if helper_user_id not in min_distances or dist < min_distances[helper_user_id]:
                min_distances[helper_user_id] = dist
            if window_start <= assignment_date <= window_end:
                assignment_counts[helper_user_id] = assignment_counts.get(helper_user_id, 0) + 1
                if helper_user_id not in task_assignment_counts:
                    task_assignment_counts[helper_user_id] = {}
                task_assignment_counts[helper_user_id][task_id] = task_assignment_counts[helper_user_id].get(task_id, 0) + 1

    # Shuffle active_users first so that brothers with the exact same metrics are randomized
    import random
    random.shuffle(active_users)
    
    min_date = date(2000, 1, 1)

    draft_assignments = []
    used_user_ids = set()
    current_assignment_map = {}
    
    # Pre-fill used_user_ids with manually assigned users
    for a in request.current_assignments:
        current_assignment_map[a.task_id] = a
        if a.main_user_id:
            used_user_ids.add(a.main_user_id)
        if a.helper_user_id:
            used_user_ids.add(a.helper_user_id)

    for task in tasks:
        allowed_roles = {m.role_id for m in matrix if m.task_id == task.id}
        
        cur_a = current_assignment_map.get(task.id)
        main_user_id = cur_a.main_user_id if cur_a else None
        helper_user_id = cur_a.helper_user_id if cur_a else None
        custom_name = cur_a.custom_name if cur_a else None

        # Find eligible main user if not manually assigned
        if not main_user_id:
            eligible_main = [
                u for u in active_users 
                if u.id not in used_user_ids and any(r.id in allowed_roles for r in u.roles)
            ]
            
            # Sort specifically to balance total workload, avoid repeating THIS task, and space out dates
            eligible_main.sort(key=lambda u: (
                assignment_counts.get(u.id, 0) + task_assignment_counts.get(u.id, {}).get(task.id, 0) * 3,
                -min_distances.get(u.id, 999),
                last_dates.get(u.id, min_date)
            ))
            
            if eligible_main:
                main_user_id = eligible_main[0].id
                used_user_ids.add(main_user_id)
            
        # Assign helper if task requires a partner and not manually assigned
        if getattr(task, 'requires_partner', False) and not helper_user_id:
            eligible_helper = [
                u for u in active_users 
                if u.id not in used_user_ids and any(r.id in allowed_roles for r in u.roles)
            ]
            
            # Sort helpers the same way
            eligible_helper.sort(key=lambda u: (
                assignment_counts.get(u.id, 0) + task_assignment_counts.get(u.id, {}).get(task.id, 0) * 3,
                -min_distances.get(u.id, 999),
                last_dates.get(u.id, min_date)
            ))
            
            if eligible_helper:
                helper_user_id = eligible_helper[0].id
                used_user_ids.add(helper_user_id)
            
        draft_assignments.append({
            "task_id": task.id,
            "main_user_id": main_user_id,
            "helper_user_id": helper_user_id,
            "custom_name": custom_name
        })

    return draft_assignments

@router.get("/export", response_class=StreamingResponse)
async def export_schedules(
    year: int,
    month: int,
    db: AsyncSession = Depends(get_db),
    current_admin: AdminUser = Depends(get_current_admin_user)
):
    # Fetch all schedules for this month
    start_date = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)

    result = await db.execute(
        select(Schedule)
        .where(Schedule.date >= start_date)
        .where(Schedule.date <= end_date)
        .options(selectinload(Schedule.assignments))
        .order_by(Schedule.date)
    )
    schedules = result.scalars().all()

    if not schedules:
        raise HTTPException(status_code=404, detail="Нет расписания на этот месяц")

    # Fetch users and tasks
    users_res = await db.execute(select(User).where(User.is_active == True))
    users = {u.id: u.full_name for u in users_res.scalars().all()}
    
    tasks_res = await db.execute(select(Task))
    tasks = {t.id: t for t in tasks_res.scalars().all()}

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Расписание {month:02d}.{year}"

    month_names = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
    month_name = month_names[month - 1]
    
    # Setup styling
    header_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = f"График служения на {month_name} {year} года"
    ws['A1'].font = header_font

    # We will build columns: A = Имя, B, C, D... = dates
    dates = [s.date for s in schedules]
    
    # Headers
    ws['A3'] = "Имя"
    ws['A3'].font = bold_font
    ws['A3'].alignment = center_align
    ws['A3'].border = thin_border
    ws.column_dimensions['A'].width = 30

    col_idx = 2
    for d in dates:
        cell = ws.cell(row=3, column=col_idx)
        cell.value = d.strftime("%d.%m")
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
        col_idx += 1

    # Aggregate user assignments (only tasks WITHOUT custom name)
    user_assignments = {} # user_id -> { date: [task_names] }
    
    for s in schedules:
        for a in s.assignments:
            task = tasks.get(a.task_id)
            if not task:
                continue
            
            # Skip tasks that require custom name — they go to PDF
            if task.requires_custom_name:
                continue
                
            task_code = task.code
            # Main user
            if a.main_user_id:
                if a.main_user_id not in user_assignments:
                    user_assignments[a.main_user_id] = {}
                if s.date not in user_assignments[a.main_user_id]:
                    user_assignments[a.main_user_id][s.date] = []
                user_assignments[a.main_user_id][s.date].append(task_code)
                
            # Helper user
            if a.helper_user_id:
                if a.helper_user_id not in user_assignments:
                    user_assignments[a.helper_user_id] = {}
                if s.date not in user_assignments[a.helper_user_id]:
                    user_assignments[a.helper_user_id][s.date] = []
                user_assignments[a.helper_user_id][s.date].append(task_code)

    # Write data rows
    row_idx = 4
    # Sort users alphabetically
    sorted_user_ids = sorted(user_assignments.keys(), key=lambda uid: users.get(uid, ""))
    
    for uid in sorted_user_ids:
        user_name = users.get(uid, f"User {uid}")
        
        # Write Name
        cell = ws.cell(row=row_idx, column=1)
        cell.value = user_name
        cell.font = bold_font
        cell.border = thin_border
        
        # Write tasks for each date
        col_idx = 2
        for d in dates:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            tasks_for_date = user_assignments[uid].get(d, [])
            if tasks_for_date:
                cell.value = ", ".join(tasks_for_date)
            col_idx += 1
            
        row_idx += 1

    # Save to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=schedule_{year}_{month:02d}.xlsx"
        }
    )

@router.get("/export-pdf", response_class=StreamingResponse)
async def export_schedules_pdf(
    year: int,
    month: int,
    db: AsyncSession = Depends(get_db),
    current_admin: AdminUser = Depends(get_current_admin_user)
):
    """Generate a beautiful minimalistic PDF chart for tasks with custom names."""
    
    start_date = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)

    result = await db.execute(
        select(Schedule)
        .where(Schedule.date >= start_date)
        .where(Schedule.date <= end_date)
        .options(selectinload(Schedule.assignments))
        .order_by(Schedule.date)
    )
    schedules = result.scalars().all()

    if not schedules:
        raise HTTPException(status_code=404, detail="Нет расписания на этот месяц")

    # Fetch users and tasks
    users_res = await db.execute(select(User))
    users = {u.id: u.full_name for u in users_res.scalars().all()}
    
    tasks_res = await db.execute(select(Task).where(Task.requires_custom_name == True).order_by(Task.order_index))
    custom_tasks = tasks_res.scalars().all()

    if not custom_tasks:
        raise HTTPException(status_code=404, detail="Нет заданий с названием")

    # Build data: rows = dates, columns = tasks
    dates = [s.date for s in schedules]
    task_ids = [t.id for t in custom_tasks]
    task_names = [t.name for t in custom_tasks]
    task_codes = [t.code for t in custom_tasks]

    # Build cell data: cell_data[date_idx][task_idx] = text
    cell_data = []
    for s in schedules:
        row = []
        for t in custom_tasks:
            # Find matching assignment
            match = None
            for a in s.assignments:
                if a.task_id == t.id:
                    match = a
                    break
            
            if match:
                parts = []
                if match.custom_name:
                    parts.append(f"«{match.custom_name}»")
                if match.main_user_id:
                    parts.append(users.get(match.main_user_id, "—"))
                row.append("\n".join(parts) if parts else "—")
            else:
                row.append("—")
        cell_data.append(row)

    # --- Build PDF with matplotlib ---
    month_names_ru = [
        'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
        'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь'
    ]
    month_name = month_names_ru[month - 1]

    n_rows = len(dates)
    n_cols = len(custom_tasks)

    # Dynamic sizing
    col_width = max(2.4, 12.0 / max(n_cols, 1))
    row_height = 0.7
    fig_width = 1.5 + n_cols * col_width
    fig_height = 2.5 + n_rows * row_height

    # Limit to reasonable page size
    fig_width = min(fig_width, 24)
    fig_height = min(fig_height, 36)

    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    ax.axis('off')
    ax.set_xlim(0, fig_width)
    ax.set_ylim(0, fig_height)

    # Colors
    bg_color = '#FAFBFC'
    header_bg = '#2D3748'
    header_text = '#FFFFFF'
    date_bg = '#EDF2F7'
    date_text = '#2D3748'
    cell_text_color = '#4A5568'
    line_color = '#E2E8F0'
    accent_color = '#667EEA'

    fig.patch.set_facecolor(bg_color)

    # Layout
    left_margin = 1.2
    top_margin = fig_height - 1.2
    table_width = fig_width - left_margin - 0.3
    cell_w = table_width / n_cols if n_cols > 0 else table_width

    # Title
    ax.text(
        fig_width / 2, fig_height - 0.4,
        f"График на {month_name} {year}",
        ha='center', va='center',
        fontsize=16, fontweight='bold',
        color=header_bg,
        family='sans-serif'
    )

    # Decorative line under title
    ax.plot(
        [fig_width * 0.3, fig_width * 0.7],
        [fig_height - 0.7, fig_height - 0.7],
        color=accent_color, linewidth=2, solid_capstyle='round'
    )

    # Header row
    header_y = top_margin - 0.5
    header_h = 0.5

    for j, task in enumerate(custom_tasks):
        x = left_margin + j * cell_w
        # Header cell background
        rect = plt.Rectangle(
            (x, header_y - header_h), cell_w, header_h,
            facecolor=header_bg, edgecolor='none',
            linewidth=0
        )
        ax.add_patch(rect)
        # Header text
        label = task.name
        if len(label) > 14:
            label = label[:12] + '…'
        ax.text(
            x + cell_w / 2, header_y - header_h / 2,
            label,
            ha='center', va='center',
            fontsize=8, fontweight='bold',
            color=header_text,
            family='sans-serif'
        )

    # Date column header
    rect = plt.Rectangle(
        (0.1, header_y - header_h), left_margin - 0.2, header_h,
        facecolor=header_bg, edgecolor='none'
    )
    ax.add_patch(rect)
    ax.text(
        (left_margin) / 2, header_y - header_h / 2,
        'Дата',
        ha='center', va='center',
        fontsize=8, fontweight='bold',
        color=header_text,
        family='sans-serif'
    )

    # Data rows
    for i, d in enumerate(dates):
        y = header_y - header_h - i * row_height
        row_bg = '#FFFFFF' if i % 2 == 0 else '#F7FAFC'

        # Date cell
        rect = plt.Rectangle(
            (0.1, y - row_height), left_margin - 0.2, row_height,
            facecolor=date_bg, edgecolor=line_color, linewidth=0.5
        )
        ax.add_patch(rect)

        weekday_names = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
        day_label = f"{d.strftime('%d.%m')} {weekday_names[d.weekday()]}"
        ax.text(
            (left_margin) / 2, y - row_height / 2,
            day_label,
            ha='center', va='center',
            fontsize=7.5, fontweight='bold',
            color=date_text,
            family='sans-serif'
        )

        # Data cells
        for j in range(n_cols):
            x = left_margin + j * cell_w
            rect = plt.Rectangle(
                (x, y - row_height), cell_w, row_height,
                facecolor=row_bg, edgecolor=line_color, linewidth=0.5
            )
            ax.add_patch(rect)

            text = cell_data[i][j]
            if text == "—":
                ax.text(
                    x + cell_w / 2, y - row_height / 2,
                    "—",
                    ha='center', va='center',
                    fontsize=7, color='#CBD5E0',
                    family='sans-serif'
                )
            else:
                # Truncate if too long
                display_text = text
                if len(display_text) > 28:
                    display_text = display_text[:26] + '…'
                ax.text(
                    x + cell_w / 2, y - row_height / 2,
                    display_text,
                    ha='center', va='center',
                    fontsize=6.5, color=cell_text_color,
                    family='sans-serif',
                    linespacing=1.3
                )

    plt.tight_layout(pad=0.5)

    # Save to PDF
    output = io.BytesIO()
    fig.savefig(output, format='pdf', bbox_inches='tight', facecolor=fig.get_facecolor(), dpi=150)
    plt.close(fig)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=chart_{year}_{month:02d}.pdf"
        }
    )
